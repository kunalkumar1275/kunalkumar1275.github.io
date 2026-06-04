"""
╔══════════════════════════════════════════════════════════════════════════════╗
║          SABR VOL SURFACE CALIBRATION — USD SWAPTIONS (β=0)                ║
║          Bachelier (Normal Vol) World                                        ║
╚══════════════════════════════════════════════════════════════════════════════╝
USAGE: python sabr_calibration_v3.py
INPUT:  sabr_input_template.xlsx (sheets: Tenors, ATM_Rates, Day1..DayN)
OUTPUT: sabr_output.xlsx (9 tabs)
"""

# ══════════════════════════════════════════════════════════════════════════════
# USER-CONFIGURABLE SETTINGS
# ══════════════════════════════════════════════════════════════════════════════
INITIAL_RHO      =  0.0
INITIAL_NU       =  0.30
RHO_LOWER        = -0.999
RHO_UPPER        =  0.999
NU_LOWER         =  0.0001
NU_UPPER         =  2.0
FTOL             =  1e-12
XTOL             =  1e-12
GTOL             =  1e-12
MAX_NFEV         =  5000
RHO_STAB_THRESH  =  0.10
NU_STAB_THRESH   =  0.05
OFFSETS_BPS      = [-300,-200,-150,-100,-50,0,+50,+100,+200,+300,+400]
NM_PRICE_THRESH  =  0.00005   # N/M threshold: price < 0.5bps of notional

# ══════════════════════════════════════════════════════════════════════════════
# IMPORTS
# ══════════════════════════════════════════════════════════════════════════════
import sys, os
import numpy as np
from scipy.optimize import least_squares, brentq
from scipy.stats import norm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

OFFSETS = np.array([o/10000.0 for o in OFFSETS_BPS])
N_STR   = len(OFFSETS)
ATM_IDX = list(OFFSETS_BPS).index(0)  # index of ATM strike

# ══════════════════════════════════════════════════════════════════════════════
# FINANCIAL FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════
def bachelier_price(F,K,T,sigma_n,is_payer=True):
    if sigma_n<1e-10: return max(F-K,0) if is_payer else max(K-F,0)
    d=(F-K)/(sigma_n*np.sqrt(T))
    if is_payer: return (F-K)*norm.cdf(d)+sigma_n*np.sqrt(T)*norm.pdf(d)
    else:        return (K-F)*norm.cdf(-d)+sigma_n*np.sqrt(T)*norm.pdf(d)

def bachelier_vega(F,K,T,sigma_n):
    if sigma_n<1e-10: return 0.0
    return np.sqrt(T)*norm.pdf((F-K)/(sigma_n*np.sqrt(T)))

def implied_normal_vol(price,F,K,T,is_payer=True):
    intrinsic=max(F-K,0) if is_payer else max(K-F,0)
    if price<=intrinsic+1e-12: return None
    try: return brentq(lambda s: bachelier_price(F,K,T,s,is_payer)-price,1e-8,0.5,xtol=1e-10)
    except: return None

def sabr_atm_correction(rho,nu,T):
    return 1.0+(2-3*rho**2)/24.0*nu**2*T

def sigma0_from_atm(sigma_mkt_atm,rho,nu,T):
    return sigma_mkt_atm/sabr_atm_correction(rho,nu,T)

def sabr_normal_vol(F,K,T,sigma0,rho,nu):
    eps=1e-8; corr=sabr_atm_correction(rho,nu,T)
    if abs(F-K)<eps: return sigma0*corr
    z=nu/sigma0*(F-K)
    denom=np.sqrt(1-2*rho*z+z**2)+z-rho
    if denom<=0 or (1-rho)<=0: return sigma0*corr
    chi=np.log(denom/(1-rho))
    return sigma0*z/chi*corr if abs(chi)>eps else sigma0*corr

# ══════════════════════════════════════════════════════════════════════════════
# INPUT SANITY CHECKS
# ══════════════════════════════════════════════════════════════════════════════
def run_input_sanity(tenors,atm_by_day,prices_by_day,n_days):
    """
    Run all input sanity checks. Returns list of check result dicts.
    Each dict: {name, passed, failures: [(day,tenor,detail),...]}
    """
    checks=[]

    # 1. All prices > 0
    failures=[]
    for d in range(n_days):
        for name,_,__ in tenors:
            for s,o in enumerate(OFFSETS):
                p=prices_by_day[d].get(name,{}).get(round(o,6),0)
                if p<=0:
                    failures.append((d+1,name,f"{int(o*10000):+d}bps price={p:.6f}"))
    checks.append({"name":"All swaption prices > 0","passed":len(failures)==0,"failures":failures})

    # 2. ATM straddle > 0
    failures=[]
    for d in range(n_days):
        for name,_,__ in tenors:
            p=prices_by_day[d].get(name,{}).get(0.0,0)
            if p<=0:
                failures.append((d+1,name,"ATM straddle price<=0"))
    checks.append({"name":"ATM straddle price > 0","passed":len(failures)==0,"failures":failures})

    # 3. Payer prices decreasing as strike increases (higher strike=more OTM payer=lower price)
    failures=[]
    payer_offsets=[o for o in OFFSETS if o>1e-6]
    for d in range(n_days):
        for name,_,__ in tenors:
            payer_prices=[prices_by_day[d].get(name,{}).get(round(o,6),0) for o in payer_offsets]
            for i in range(len(payer_prices)-1):
                if payer_prices[i]<payer_prices[i+1]-1e-8:
                    failures.append((d+1,name,
                        f"Payer: {int(payer_offsets[i]*10000):+d}bps({payer_prices[i]:.6f})"
                        f"<{int(payer_offsets[i+1]*10000):+d}bps({payer_prices[i+1]:.6f})"))
    checks.append({"name":"Payer prices decrease as strike increases","passed":len(failures)==0,"failures":failures})

    # 4. Receiver prices increase as strike increases (higher strike=more ITM receiver=higher price)
    failures=[]
    recv_offsets=[o for o in OFFSETS if o<-1e-6]
    for d in range(n_days):
        for name,_,__ in tenors:
            recv_prices=[prices_by_day[d].get(name,{}).get(round(o,6),0) for o in recv_offsets]
            for i in range(len(recv_prices)-1):
                if recv_prices[i]>recv_prices[i+1]+1e-8:
                    failures.append((d+1,name,
                        f"Recv: {int(recv_offsets[i]*10000):+d}bps({recv_prices[i]:.6f})"
                        f">{int(recv_offsets[i+1]*10000):+d}bps({recv_prices[i+1]:.6f})"))
    checks.append({"name":"Receiver prices increase as strike increases","passed":len(failures)==0,"failures":failures})

    # 5. ATM forward rates reasonable (between -5% and 20%)
    failures=[]
    for d in range(n_days):
        for name,_,__ in tenors:
            F=atm_by_day[d].get(name,0)
            if not (-0.05<=F<=0.20):
                failures.append((d+1,name,f"F={F:.4%} outside [-5%,20%]"))
    checks.append({"name":"ATM forward rates in [-5%, 20%]","passed":len(failures)==0,"failures":failures})

    # 6. ATM straddle > first payer price (+50bps)
    # Straddle = ATM_payer + ATM_receiver. First payer is OTM so cheaper than ATM payer.
    # Therefore straddle must always exceed any individual OTM payer price.
    failures=[]
    for d in range(n_days):
        for name,_,__ in tenors:
            prices=prices_by_day[d].get(name,{})
            atm_p=prices.get(0.0,0)
            first_payer=prices.get(round(OFFSETS[ATM_IDX+1],6),0)
            if atm_p>0 and first_payer>0:
                if atm_p<=first_payer:
                    failures.append((d+1,name,
                        f"Straddle={atm_p:.6f} <= first payer(+{int(OFFSETS[ATM_IDX+1]*10000)}bps)={first_payer:.6f}"))
    checks.append({"name":"ATM straddle > first payer price (+50bps) [arbitrage check]",
                   "passed":len(failures)==0,"failures":failures})

    # 7. ATM straddle > first receiver price (-50bps)
    # Same logic: straddle = ATM_payer + ATM_receiver > any individual OTM receiver price.
    failures=[]
    for d in range(n_days):
        for name,_,__ in tenors:
            prices=prices_by_day[d].get(name,{})
            atm_p=prices.get(0.0,0)
            first_recv=prices.get(round(OFFSETS[ATM_IDX-1],6),0)
            if atm_p>0 and first_recv>0:
                if atm_p<=first_recv:
                    failures.append((d+1,name,
                        f"Straddle={atm_p:.6f} <= first receiver({int(OFFSETS[ATM_IDX-1]*10000)}bps)={first_recv:.6f}"))
    checks.append({"name":"ATM straddle > first receiver price (-50bps) [arbitrage check]",
                   "passed":len(failures)==0,"failures":failures})

    return checks

# ══════════════════════════════════════════════════════════════════════════════
# CALIBRATION
# ══════════════════════════════════════════════════════════════════════════════
def calibrate_tenor(prices_dict,F,T_exp,x0_rho,x0_nu):
    straddle_p=prices_dict.get(0.0,prices_dict.get(0,None))
    if straddle_p is None: raise ValueError("ATM straddle price not found")
    sigma_mkt_atm=straddle_p/(2*np.sqrt(T_exp)*norm.pdf(0))
    strikes=F+OFFSETS
    market_vols=[]
    for i,offset in enumerate(OFFSETS):
        K=strikes[i]
        if abs(offset)<1e-6:
            mv=sigma_mkt_atm
        else:
            p=prices_dict.get(round(offset,6),None)
            if p is None:
                print(f"    Warning: price missing {offset*10000:+.0f}bps"); mv=sigma_mkt_atm
            else:
                mv=implied_normal_vol(p,F,K,T_exp,offset>0)
                if mv is None:
                    print(f"    Warning: Brent failed {offset*10000:+.0f}bps"); mv=sigma_mkt_atm
        market_vols.append(mv)
    market_vols=np.array(market_vols)
    vegas=np.array([bachelier_vega(F,K,T_exp,mv) for K,mv in zip(strikes,market_vols)])

    conv_status={"code":4,"message":"Max iterations reached"}
    def residuals(params):
        rho,nu=params
        s0=sigma0_from_atm(sigma_mkt_atm,rho,nu,T_exp)
        mdl=np.array([sabr_normal_vol(F,K,T_exp,s0,rho,nu) for K in strikes])
        return np.sqrt(vegas)*(mdl-market_vols)

    res=least_squares(residuals,x0=[x0_rho,x0_nu],
                      bounds=([RHO_LOWER,NU_LOWER],[RHO_UPPER,NU_UPPER]),
                      method='trf',ftol=FTOL,xtol=XTOL,gtol=GTOL,max_nfev=MAX_NFEV)
    rho_c,nu_c=res.x
    # Determine convergence code
    if res.status==1:   conv_status={"code":1,"message":"ftol: cost J converged"}
    elif res.status==2: conv_status={"code":2,"message":"xtol: parameters converged"}
    elif res.status==3: conv_status={"code":3,"message":"gtol: gradient flat"}
    elif res.status==4: conv_status={"code":4,"message":"Max iterations reached"}
    else:               conv_status={"code":res.status,"message":res.message}

    s0_c=sigma0_from_atm(sigma_mkt_atm,rho_c,nu_c,T_exp)
    mdl_ivs=np.array([sabr_normal_vol(F,K,T_exp,s0_c,rho_c,nu_c) for K in strikes])

    # Re-generate prices for round-trip
    regen_prices=[]
    for i,offset in enumerate(OFFSETS):
        K=strikes[i]
        if abs(offset)<1e-6:
            p=2*mdl_ivs[i]*np.sqrt(T_exp)*norm.pdf(0)
        else:
            p=bachelier_price(F,K,T_exp,mdl_ivs[i],offset>0)
        regen_prices.append(p)

    input_prices=np.array([prices_dict.get(round(o,6),0) for o in OFFSETS])

    return {
        "sigma_mkt_atm": sigma_mkt_atm,
        "sigma0":        s0_c,
        "rho":           rho_c,
        "nu":            nu_c,
        "correction":    sabr_atm_correction(rho_c,nu_c,T_exp),
        "cost_J":        float(np.sum(vegas*(mdl_ivs-market_vols)**2)),
        "nfev":          res.nfev,
        "conv_code":     conv_status["code"],
        "conv_msg":      conv_status["message"],
        "mkt_ivs":       market_vols,
        "mdl_ivs":       mdl_ivs,
        "strikes":       strikes,
        "vegas":         vegas,
        "input_prices":  input_prices,
        "regen_prices":  np.array(regen_prices),
        "F":             F,
        "T_exp":         T_exp,
    }

def run_calibration(tenors,atm_by_day,prices_by_day,n_days):
    results=[]; prev={t[0]:(INITIAL_RHO,INITIAL_NU) for t in tenors}
    for d in range(n_days):
        print(f"\n  Day {d+1}:")
        day={}
        for name,T_exp,_ in tenors:
            F=atm_by_day[d][name]; rho0,nu0=prev[name]
            try:
                r=calibrate_tenor(prices_by_day[d][name],F,T_exp,rho0,nu0)
                print(f"    {name:8s}: σ_ATM={r['sigma_mkt_atm']*10000:.2f}bps "
                      f"ρ={r['rho']:+.4f} ν={r['nu']:.4f} "
                      f"ATM_err={(r['mdl_ivs'][ATM_IDX]-r['mkt_ivs'][ATM_IDX])*10000:.4f}bps "
                      f"code={r['conv_code']}")
                day[name]=r; prev[name]=(r['rho'],r['nu'])
            except Exception as e:
                print(f"    {name:8s}: FAILED — {e}"); day[name]=None
        results.append(day)
    return results

# ══════════════════════════════════════════════════════════════════════════════
# INPUT READING
# ══════════════════════════════════════════════════════════════════════════════
def read_input(filepath,n_days):
    print(f"\nReading: {filepath}")
    wb=openpyxl.load_workbook(filepath,data_only=True)
    ws_t=wb["Tenors"]; tenors=[]
    for row in ws_t.iter_rows(min_row=2,values_only=True):
        if row[0] is None: break
        tenors.append((str(row[0]).strip(),float(row[1]),float(row[2])))
    ws_r=wb["ATM_Rates"]; atm_by_day=[{} for _ in range(n_days)]
    for row in ws_r.iter_rows(min_row=2,values_only=True):
        if row[0] is None: break
        name=str(row[0]).strip()
        for d in range(n_days): atm_by_day[d][name]=float(row[1+d])
    prices_by_day=[]
    for d in range(n_days):
        sname=f"Day{d+1}"
        if sname not in wb.sheetnames: raise ValueError(f"Sheet '{sname}' not found")
        ws_p=wb[sname]; day_prices={}
        for row in ws_p.iter_rows(min_row=2,values_only=True):
            if row[0] is None: break
            name=str(row[0]).strip()
            day_prices[name]={round(OFFSETS[s],6):float(row[1+s]) if row[1+s] else 0.0
                              for s in range(N_STR)}
        prices_by_day.append(day_prices)
    print(f"  Loaded: {[t[0] for t in tenors]}, {n_days} days")
    return tenors,atm_by_day,prices_by_day

# ══════════════════════════════════════════════════════════════════════════════
# STYLE HELPERS
# ══════════════════════════════════════════════════════════════════════════════
HDR_NAVY="FF1F3864"; HDR_GREEN="FF375623"; HDR_TEAL="FF1F6B75"
WHITE="FFFFFFFF"; BLACK_="FF000000"; BLUE_="FF0000FF"
ALT="FFD9E1F2"; GREEN_BG="FFE2EFDA"; AMBER_BG="FFFFF2CC"
PASS_BG="FF92D050"; FAIL_BG="FFFF0000"; BLANK_BG="FFD0D0D0"
WARN_BG="FFFF6600"; NM_BG="FFD0D0D0"

_thin=Side(style='thin',color='FF000000')
_BDR=Border(left=_thin,right=_thin,top=_thin,bottom=_thin)

def _s(ws,r,c,v,fg=BLACK_,bg=None,bold=False,fmt=None,align='center',wrap=False,sz=9,italic=False):
    if r<1 or c<1: return
    cell=ws.cell(row=r,column=c,value=v)
    cell.font=Font(name='Arial',color=fg,bold=bold,size=sz,italic=italic)
    cell.fill=PatternFill('solid',start_color=bg) if bg else PatternFill()
    cell.border=_BDR
    cell.alignment=Alignment(horizontal=align,vertical='center',wrap_text=wrap)
    if fmt: cell.number_format=fmt
    return cell

def _h(ws,r,c,v,bg=HDR_NAVY,fg=WHITE,sz=9,wrap=False,bold=True):
    return _s(ws,r,c,v,fg=fg,bg=bg,bold=bold,align='center',sz=sz,wrap=wrap)

def _c(ws,r,c,v,fmt=None,bg=None): return _s(ws,r,c,v,fmt=fmt,bg=bg,align='center')
def _i(ws,r,c,v,fmt=None,bg=None): return _s(ws,r,c,v,fg=BLUE_,fmt=fmt,bg=bg,align='center')

def _note(ws,r,c,v,end_col,bg=AMBER_BG):
    _s(ws,r,c,v,fg=BLACK_,bg=bg,italic=True,align='left',wrap=True,sz=9)
    if end_col>c: ws.merge_cells(f'{get_column_letter(c)}{r}:{get_column_letter(end_col)}{r}')

def _blank_row(ws,r,n_cols):
    for c in range(1,n_cols+2):
        cell=ws.cell(row=r,column=c,value="")
        cell.fill=PatternFill('solid',start_color=BLANK_BG)
        cell.border=Border()

def _offset_label(o): return "ATM" if abs(o)<1e-6 else f"{int(o*10000):+d}"

def _write_legend(ws, start_row, start_col, title, items, col_widths=(10,42)):
    """
    Write a colour legend block on the right side of a sheet.
    items = list of (bg_color, fg_color, label, description)
    """
    lc=start_col; dc=start_col+1
    ws.column_dimensions[get_column_letter(lc)].width=col_widths[0]
    ws.column_dimensions[get_column_letter(dc)].width=col_widths[1]
    _h(ws,start_row,lc,title,bg=HDR_NAVY)
    ws.merge_cells(f'{get_column_letter(lc)}{start_row}:{get_column_letter(dc)}{start_row}')
    row=start_row+1
    for bg,fg,lbl,desc in items:
        _s(ws,row,lc,lbl,bold=True,fg=fg,bg=bg,align='center')
        _s(ws,row,dc,desc,align='left',wrap=True,fg=BLACK_)
        ws.row_dimensions[row].height=max(15,min(50,len(desc)//2))
        row+=1
    return row


def _set_grid_widths(ws,n_strikes,col_start=1):
    ws.column_dimensions[get_column_letter(col_start)].width=9
    for s in range(n_strikes):
        ws.column_dimensions[get_column_letter(col_start+1+s)].width=9

def write_stacked_grid(ws,results,tenors,n_days,val_fn,fmt,bg_fn=None,col_start=1,day_bg=HDR_TEAL):
    row=2
    for d in range(n_days):
        end_col=col_start+N_STR
        _h(ws,row,col_start,f"DAY {d+1}",bg=day_bg)
        ws.merge_cells(f'{get_column_letter(col_start)}{row}:{get_column_letter(end_col)}{row}')
        row+=1
        _h(ws,row,col_start,"Tenor",bg=HDR_NAVY)
        for s,o in enumerate(OFFSETS): _h(ws,row,col_start+1+s,_offset_label(o),bg=HDR_NAVY)
        row+=1
        for i,(name,_,__) in enumerate(tenors):
            bg=ALT if i%2==0 else None
            _s(ws,row,col_start,name,bold=True,bg=bg)
            r=results[d].get(name)
            for s,o in enumerate(OFFSETS):
                val=val_fn(r,s) if r else 0
                is_atm=abs(o)<1e-6
                cell_bg=AMBER_BG if is_atm else (bg_fn(val) if bg_fn else bg)
                _c(ws,row,col_start+1+s,val,fmt=fmt,bg=cell_bg)
            row+=1
        _blank_row(ws,row,end_col); row+=1
    return row

# ══════════════════════════════════════════════════════════════════════════════
# OUTPUT WRITER
# ══════════════════════════════════════════════════════════════════════════════
def write_output(results,tenors,atm_by_day,prices_by_day,n_days,sanity_checks,out_path):
    N_TEN=len(tenors)
    wb=openpyxl.Workbook()

    # ── 1. Overview ──────────────────────────────────────────────────────────
    ws=wb.active; ws.title="1_Overview"
    ws.column_dimensions['A'].width=34; ws.column_dimensions['B'].width=74
    _h(ws,1,1,"SABR Vol Surface Calibration — USD Swaptions (β=0, Bachelier World)",sz=12)
    ws.merge_cells('A1:B1')
    ws['A1'].font=Font(name='Arial',bold=True,color=WHITE,size=12)
    rows=[
        ("── MODEL ──","",HDR_NAVY),
        ("Model","SABR Hagan 2002, β=0 → Normal (Bachelier) vol backbone",None),
        ("Why β=0","Handles negative rates. β=1 (lognormal) breaks at K≤0.",None),
        ("Vol units","Normal vol σ_N in bps. Market standard for USD swaptions post-2008.",None),
        ("Prices","Totem prices as fraction of notional (annuity A=1). No discount curve needed.",None),
        ("── BACHELIER FORMULA ──","",HDR_TEAL),
        ("Payer","P=(F−K)·N(d)+σ_N·√T·φ(d)   d=(F−K)/(σ_N·√T)",None),
        ("Receiver","P=(K−F)·N(−d)+σ_N·√T·φ(d)",None),
        ("ATM Straddle","Straddle=2·σ_N·√T·φ(0)   [d=0 at K=F, payer+receiver]",None),
        ("φ(0)","φ(0)=1/√(2π)≈0.3989. Height of standard normal PDF at peak.",None),
        ("IV constraint","σ_N>0. No singularity. Handles negative rates.",None),
        ("── CALIBRATION FLOW ──","",HDR_GREEN),
        ("Step 1","σ_mkt_ATM=Straddle/(2·√T·φ(0)). Direct, no iteration.",None),
        ("Step 2","Brent root-finding on Bachelier → σ_mkt(Kᵢ) for non-ATM strikes.",None),
        ("Step 3 each L-M iter","σ₀(ρ,ν)=σ_mkt_ATM/[1+(2−3ρ²)·ν²·T/24] → ATM error=0 by construction.",None),
        ("Step 4","L-M minimises J(ρ,ν). σ₀ derived each iter. Warm-start from prev day.",None),
        ("── COST FUNCTION ──","",HDR_GREEN),
        ("J(ρ,ν)","J=Σᵢ vegaᵢ·[σ_SABR(Kᵢ)−σ_mkt(Kᵢ)]²   (vega-weighted IV SSE)",None),
        ("⚠ IV not price","vega×(IV diff)², NOT vega×(price diff)². Gives wings enough weight to identify ρ,ν.",None),
        ("Vega","vegaᵢ=√T·φ(dᵢ) at MARKET IV. Fixed throughout — no circular dependency.",None),
        ("── STOPPING CRITERIA ──","",HDR_TEAL),
        ("Code 1","ftol: relative change in cost J < 1e-12 → CONVERGED",None),
        ("Code 2","xtol: relative change in {ρ,ν} < 1e-12 → CONVERGED",None),
        ("Code 3","gtol: gradient of J < 1e-12 → CONVERGED",None),
        ("Code 4",f"max_nfev={MAX_NFEV} reached without convergence → INVESTIGATE",None),
        ("── SABR FORMULA β=0 ──","",HDR_GREEN),
        ("ATM","σ_SABR(ATM)=σ₀·[1+(2−3ρ²)·ν²·T/24]  ← set=σ_mkt_ATM each iteration",None),
        ("Away","σ_SABR(K)=σ₀·(z/χ(z))·[1+(2−3ρ²)·ν²·T/24]",None),
        ("z / χ(z)","z=ν/σ₀·(F−K)   χ(z)=ln[(√(1−2ρz+z²)+z−ρ)/(1−ρ)]",None),
        ("── PARAMETERS ──","",HDR_GREEN),
        ("σ_mkt_ATM","Market ATM vol from straddle. Tracks market — expected to drift daily.",None),
        ("σ₀","SABR backbone=σ_mkt_ATM/correction. Slightly below σ_mkt_ATM.",None),
        ("ρ","Skew. ρ<0→lower strikes higher vol. Typical USD: −0.30 to −0.10. Should be STABLE.",None),
        ("ν","Curvature. ν>0. Wing height. Typical USD: 0.25−0.50. Should be STABLE.",None),
        ("⚠ Independence","Each tenor calibrated independently. Parameters do NOT affect other tenors.",None),
        ("── PRICE ROUND-TRIP ──","",HDR_GREEN),
        ("N/M threshold",f"% error shown as N/M when input price < {NM_PRICE_THRESH*10000:.1f}bps of notional (option nearly worthless).",None),
        ("Good fit","Near-ATM |bps err|<0.1 or |%err|<0.5%. Wings: larger errors acceptable (vega≈0).",None),
        ("── USER SETTINGS ──","",HDR_TEAL),
        ("Initial ρ/ν",f"ρ={INITIAL_RHO}, ν={INITIAL_NU}   (edit at top of script)",None),
        ("Bounds",f"ρ∈[{RHO_LOWER},{RHO_UPPER}]  ν∈[{NU_LOWER},{NU_UPPER}]",None),
        (f"Stability","ρ PASS if range<{RHO_STAB_THRESH}. ν PASS if range<{NU_STAB_THRESH}.",None),
    ]
    for i,(k,v,sbg) in enumerate(rows,3):
        if sbg:
            _h(ws,i,1,k,bg=sbg); ws.merge_cells(f'A{i}:B{i}')
        else:
            amber="⚠" in k or "⚠" in v
            _s(ws,i,1,k,bold=True,align='left',bg=GREEN_BG if amber else None)
            _s(ws,i,2,v,align='left',wrap=True,bg=GREEN_BG if amber else None)
        ws.row_dimensions[i].height=26 if v and len(v)>90 else 15
    ws.freeze_panes="A3"

    # ── 2. Input Sanity Checks ───────────────────────────────────────────────
    ws=wb.create_sheet("2_Input_Sanity")
    ws.column_dimensions['A'].width=45
    ws.column_dimensions['B'].width=12
    ws.column_dimensions['C'].width=70
    _h(ws,1,1,"INPUT SANITY CHECKS — Run before calibration on raw Totem prices",bg=HDR_NAVY,sz=10)
    ws.merge_cells('A1:C1')
    _h(ws,2,1,"Check Description"); _h(ws,2,2,"Result"); _h(ws,2,3,"Failures (Day | Tenor | Detail)")
    for i,chk in enumerate(sanity_checks,3):
        bg=ALT if i%2==1 else None
        _s(ws,i,1,chk['name'],align='left',bg=bg)
        if chk['passed']:
            _s(ws,i,2,"PASS",bold=True,fg=WHITE,bg=PASS_BG,align='center')
            _s(ws,i,3,"—",align='center',bg=bg)
        else:
            _s(ws,i,2,"FAIL",bold=True,fg=WHITE,bg=FAIL_BG,align='center')
            failure_text=" | ".join([f"D{d} {ten} {det}" for d,ten,det in chk['failures'][:10]])
            if len(chk['failures'])>10: failure_text+=f" ... (+{len(chk['failures'])-10} more)"
            _s(ws,i,3,failure_text,align='left',wrap=True,bg=WARN_BG)
        ws.row_dimensions[i].height=20
    note_row=len(sanity_checks)+4
    _note(ws,note_row,1,
          "Checks run on raw Totem input prices before calibration. "
          "FAIL = investigate input data before trusting calibration results. "
          "Payer prices should decrease left→right (higher strike=more OTM). "
          "Receiver prices should increase left→right (higher strike=more ITM).",3)
    ws.freeze_panes="A3"

    # ── 3. Input Prices ──────────────────────────────────────────────────────
    ws=wb.create_sheet("3_Input_Prices")
    _set_grid_widths(ws,N_STR)
    _h(ws,1,1,f"INPUT: Totem Swaption Prices (fraction of notional, A=1) — {n_days} Days",bg=HDR_NAVY,sz=10,wrap=True)
    ws.merge_cells(f'A1:{get_column_letter(N_STR+1)}1')
    ws.row_dimensions[1].height=28

    row=2
    for d in range(n_days):
        _h(ws,row,1,f"DAY {d+1}",bg=HDR_TEAL)
        ws.merge_cells(f'A{row}:{get_column_letter(N_STR+1)}{row}'); row+=1
        _h(ws,row,1,"Tenor",bg=HDR_NAVY)
        for s,o in enumerate(OFFSETS): _h(ws,row,2+s,_offset_label(o),bg=HDR_NAVY)
        row+=1
        for i,(name,_,__) in enumerate(tenors):
            bg=ALT if i%2==0 else None
            _s(ws,row,1,name,bold=True,bg=bg)
            for s,o in enumerate(OFFSETS):
                p=prices_by_day[d].get(name,{}).get(round(o,6),0)
                _i(ws,row,2+s,round(p,7),fmt='0.0000000',bg=AMBER_BG if abs(o)<1e-6 else bg)
            row+=1
        _blank_row(ws,row,N_STR+2); row+=1

    # ATM rates
    _h(ws,row,1,"ATM Forward Swap Rates",bg=HDR_GREEN)
    ws.merge_cells(f'A{row}:{get_column_letter(n_days+1)}{row}'); row+=1
    _h(ws,row,1,"Tenor")
    for d in range(n_days): _h(ws,row,2+d,f"Day {d+1}")
    row+=1
    for i,(name,_,__) in enumerate(tenors):
        bg=ALT if i%2==0 else None
        _s(ws,row,1,name,bold=True,bg=bg)
        for d in range(n_days): _i(ws,row,2+d,atm_by_day[d].get(name,0),fmt='0.000%',bg=bg)
        row+=1
    # Legend
    _write_legend(ws, 2, N_STR+3, "COLOUR LEGEND",
        [(AMBER_BG, BLACK_, "Amber", "ATM straddle price (K=F, payer+receiver combined). Fraction of notional, A=1. Inverted directly: σ_mkt_ATM=Straddle/(2·√T·φ(0))."),
         (BLUE_,    BLACK_, "Blue",  "Non-ATM payer (K>F) or receiver (K<F) prices. Same convention: fraction of notional, A=1. Inverted via Brent method on Bachelier formula."),
        ], col_widths=(10,40))
    ws.freeze_panes="A2"

    # ── 4. Market IVs ────────────────────────────────────────────────────────
    ws=wb.create_sheet("4_Market_IVs_bps")
    _set_grid_widths(ws,N_STR)
    _h(ws,1,1,"DERIVED: Market Implied Normal Vols (bps). ATM=σ_mkt_ATM=Straddle/(2·√T·φ(0)) direct inversion. Non-ATM=Brent inversion of Bachelier formula on input prices. These are the calibration targets.",bg=HDR_NAVY,sz=9,wrap=True)
    ws.merge_cells(f'A1:{get_column_letter(N_STR+1)}1'); ws.row_dimensions[1].height=30
    write_stacked_grid(ws,results,tenors,n_days,
                       val_fn=lambda r,s: round(r['mkt_ivs'][s]*10000,2),fmt='0.00')
    # Legend
    _write_legend(ws, 2, N_STR+3, "COLOUR LEGEND",
        [(AMBER_BG, BLACK_, "Amber", "ATM strike (K=F). IV implied directly from straddle: σ_mkt_ATM=Straddle/(2·√T·φ(0)). Zero error by construction in model."),
         (ALT,      BLACK_, "Blue alt", "Non-ATM market IVs from Brent inversion of Bachelier formula. These are the calibration targets for L-M optimisation."),
        ], col_widths=(10,40))
    ws.freeze_panes="A2"

    # ── 5. Vega Weights ──────────────────────────────────────────────────────
    ws=wb.create_sheet("5_Vega_Weights")
    _set_grid_widths(ws,N_STR)
    _h(ws,1,1,"DERIVED: Bachelier Vega=√T·φ(d) at Market IV. Cost function weight per strike. Computed at market IV and held fixed throughout L-M. ATM has highest vega=φ(0)·√T. Deep OTM→vega≈0→near-zero weight in calibration.",bg=HDR_NAVY,sz=9,wrap=True)
    ws.merge_cells(f'A1:{get_column_letter(N_STR+1)}1'); ws.row_dimensions[1].height=30
    def vbg(v): return GREEN_BG if v>0.3 else (AMBER_BG if v>0.05 else None)
    write_stacked_grid(ws,results,tenors,n_days,
                       val_fn=lambda r,s: round(r['vegas'][s],4),fmt='0.0000',bg_fn=vbg)
    # Legend
    _write_legend(ws, 2, N_STR+3, "COLOUR LEGEND",
        [(GREEN_BG, BLACK_, "Green",  "High vega > 0.30. Near-ATM strikes. These carry the most weight in the L-M cost function. SABR fits these strikes most precisely."),
         (AMBER_BG, BLACK_, "Amber",  "ATM strike (always highest vega = φ(0)·√T ≈ 0.399·√T). Also used for mid-range vega 0.05–0.30."),
         (None,     BLACK_, "White",  "Near-zero vega < 0.05. Deep OTM/ITM strikes. Effectively ignored by the optimiser. Large IV errors here have minimal price impact."),
        ], col_widths=(10,40))
    ws.freeze_panes="A2"

    # ── 6. Calibrated Parameters ─────────────────────────────────────────────
    ws=wb.create_sheet("6_Calibrated_Params")
    ws.column_dimensions['A'].width=12
    for c in range(2,n_days*5+5): ws.column_dimensions[get_column_letter(c)].width=11
    _h(ws,1,1,f"Calibrated SABR Parameters — {N_TEN} Tenors × {n_days} Days",bg=HDR_NAVY)
    ws.merge_cells(f'A1:{get_column_letter(1+n_days*5)}1')
    _h(ws,2,1,"Tenor")
    col=2
    for d in range(n_days):
        _h(ws,2,col,f"D{d+1} σ_mkt_ATM(bps)"); _h(ws,2,col+1,f"D{d+1} σ₀(bps)")
        _h(ws,2,col+2,f"D{d+1} ρ"); _h(ws,2,col+3,f"D{d+1} ν"); _h(ws,2,col+4,f"D{d+1} Cost J")
        col+=5
    for i,(name,_,__) in enumerate(tenors,3):
        bg=ALT if i%2==1 else None
        _s(ws,i,1,name,bold=True,bg=bg); col=2
        for d in range(n_days):
            r=results[d].get(name)
            _c(ws,i,col,  round(r['sigma_mkt_atm']*10000,2) if r else "ERR",fmt='0.00',bg=AMBER_BG)
            _c(ws,i,col+1,round(r['sigma0']*10000,2) if r else "ERR",fmt='0.00',bg=bg)
            _c(ws,i,col+2,round(r['rho'],4) if r else "ERR",fmt='0.0000',bg=bg)
            _c(ws,i,col+3,round(r['nu'],4) if r else "ERR",fmt='0.0000',bg=bg)
            _c(ws,i,col+4,r['cost_J'] if r else "ERR",fmt='0.00E+00',bg=bg)
            col+=5
    # Stability
    r0=N_TEN+5
    _h(ws,r0,1,"TEST 1 — Stability of {ρ,ν} across Days",bg=HDR_GREEN)
    ws.merge_cells(f'A{r0}:K{r0}'); r0+=1
    for lbl,c in [("Tenor",1),("σ_ATM D1",2),("σ_ATM Last",3),("Δσ_ATM",4),
                   ("ρ min",5),("ρ max",6),("ρ range",7),("ρ PASS?",8),
                   ("ν min",9),("ν max",10),("ν PASS?",11)]:
        _h(ws,r0,c,lbl)
    for i,(name,_,__) in enumerate(tenors,r0+1):
        bg=ALT if i%2==1 else None
        valid=[results[d][name] for d in range(n_days) if results[d].get(name)]
        _s(ws,i,1,name,bold=True,bg=bg)
        if not valid: continue
        atms=[r['sigma_mkt_atm']*10000 for r in valid]
        rhos=[r['rho'] for r in valid]; nus=[r['nu'] for r in valid]
        rr=max(rhos)-min(rhos); nr=max(nus)-min(nus)
        rp=rr<RHO_STAB_THRESH; np_=nr<NU_STAB_THRESH
        _c(ws,i,2,round(atms[0],2),fmt='0.00',bg=AMBER_BG)
        _c(ws,i,3,round(atms[-1],2),fmt='0.00',bg=AMBER_BG)
        _c(ws,i,4,round(atms[-1]-atms[0],2),fmt='+0.00;-0.00;0.00',bg=AMBER_BG)
        _c(ws,i,5,round(min(rhos),4),fmt='0.0000',bg=bg)
        _c(ws,i,6,round(max(rhos),4),fmt='0.0000',bg=bg)
        _c(ws,i,7,round(rr,4),fmt='0.0000',bg=bg)
        _s(ws,i,8,"PASS" if rp else "FAIL",align='center',bold=True,fg=WHITE,bg=PASS_BG if rp else FAIL_BG)
        _c(ws,i,9,round(min(nus),4),fmt='0.0000',bg=bg)
        _c(ws,i,10,round(max(nus),4),fmt='0.0000',bg=bg)
        _s(ws,i,11,"PASS" if np_ else "FAIL",align='center',bold=True,fg=WHITE,bg=PASS_BG if np_ else FAIL_BG)
    _note(ws,r0+N_TEN+2,1,f"Δσ_ATM=expected drift no pass/fail. ρ PASS<{RHO_STAB_THRESH}. ν PASS<{NU_STAB_THRESH}.",11)
    # Legend for amber ATM column
    leg_row=3
    _write_legend(ws, leg_row, 1+n_days*5+2, "COLOUR LEGEND",
        [(AMBER_BG, BLACK_, "Amber",  f"σ_mkt_ATM column — the market ATM normal vol (bps) implied directly from the ATM straddle price each day. Expected to drift as market moves. No pass/fail applied. σ₀ (SABR backbone) is slightly lower: σ₀=σ_mkt_ATM/correction where correction=[1+(2-3ρ²)·ν²·T/24]>1."),
         (PASS_BG,  WHITE,  "Green",  "PASS — parameter range across days is within stability threshold (ρ<{RHO_STAB_THRESH}, ν<{NU_STAB_THRESH}). Smile shape is stable."),
         (FAIL_BG,  WHITE,  "Red",    "FAIL — parameter range exceeds threshold. Investigate: possible regime change, input data error, or calibration instability (warm-start from different local minimum)."),
        ], col_widths=(10,45))
    ws.freeze_panes="B3"

    # ── 7. Calibration Quality ───────────────────────────────────────────────
    ws=wb.create_sheet("7_Calib_Quality")
    ws.column_dimensions['A'].width=12
    for c in range(2,n_days+3): ws.column_dimensions[get_column_letter(c)].width=11
    ws.column_dimensions[get_column_letter(n_days+3)].width=3
    ws.column_dimensions[get_column_letter(n_days+4)].width=10
    ws.column_dimensions[get_column_letter(n_days+5)].width=35
    ws.column_dimensions[get_column_letter(n_days+6)].width=55

    _h(ws,1,1,"CALIBRATION QUALITY — Convergence codes and output sanity checks per tenor per day",bg=HDR_NAVY,sz=10,wrap=True)
    ws.merge_cells(f'A1:{get_column_letter(n_days+6)}1'); ws.row_dimensions[1].height=28

    # Convergence grid
    _h(ws,2,1,"CONVERGENCE CODE — How L-M optimisation ended for each tenor/day",bg=HDR_TEAL)
    ws.merge_cells(f'A2:{get_column_letter(n_days+1)}2')
    _h(ws,2,n_days+4,"Code",bg=HDR_TEAL)
    _h(ws,2,n_days+5,"Short meaning",bg=HDR_TEAL)
    _h(ws,2,n_days+6,"Detailed explanation",bg=HDR_TEAL)

    _h(ws,3,1,"Tenor")
    for d in range(n_days): _h(ws,3,2+d,f"Day {d+1}")

    # Code legend — expanded
    legend=[
        (1,"ftol: cost J converged ✅",PASS_BG,
         "The relative change in the cost function J between successive iterations fell below ftol=1e-12. "
         "This means J is no longer improving meaningfully — the optimiser has found a minimum. "
         "Most common convergence reason. GOOD."),
        (2,"xtol: parameters converged ✅",PASS_BG,
         "The relative change in the parameter vector {ρ, ν} between iterations fell below xtol=1e-12. "
         "Parameters are no longer moving — they have settled at stable values. "
         "Indicates the solution is well-defined and stable. GOOD."),
        (3,"gtol: gradient flat ✅",PASS_BG,
         "The norm of the gradient of J with respect to {ρ, ν} fell below gtol=1e-12. "
         "A flat gradient means we are at a stationary point (minimum). "
         "Can occasionally indicate a saddle point rather than a true minimum, but rare for this 2-parameter problem. GOOD."),
        (4,f"Max iterations ❌ INVESTIGATE",FAIL_BG,
         f"The optimiser reached the hard cap of {MAX_NFEV} function evaluations without satisfying any of the "
         "convergence criteria above. This means the cost function did not converge. Possible causes: "
         "(1) initial guess {ρ₀, ν₀} far from true values — try different starting point; "
         "(2) input prices have errors or arbitrage — check Tab 2_Input_Sanity; "
         "(3) SABR model cannot fit this smile shape — check market prices for that tenor."),
    ]
    for li,(code,short,lbg,detail) in enumerate(legend,3):
        _s(ws,li,n_days+4,code,bold=True,fg=WHITE,bg=lbg,align='center')
        _s(ws,li,n_days+5,short,align='left',bg=lbg if code==4 else None,
           fg=WHITE if code==4 else BLACK_)
        _s(ws,li,n_days+6,detail,align='left',wrap=True,
           bg=FAIL_BG if code==4 else GREEN_BG,
           fg=WHITE if code==4 else BLACK_)
        ws.row_dimensions[li].height=55

    for i,(name,_,__) in enumerate(tenors,4):
        bg=ALT if i%2==0 else None
        _s(ws,i,1,name,bold=True,bg=bg)
        for d in range(n_days):
            r=results[d].get(name)
            code=r['conv_code'] if r else 4
            cbg=PASS_BG if code in [1,2,3] else FAIL_BG
            _s(ws,i,2+d,code,bold=True,fg=WHITE,bg=cbg,align='center')

    # Output sanity checks
    r0=N_TEN+6
    _h(ws,r0,1,"OUTPUT SANITY CHECKS — Post-calibration quality checks",bg=HDR_GREEN)
    ws.merge_cells(f'A{r0}:{get_column_letter(n_days+6)}{r0}'); r0+=1

    # Check headers
    _h(ws,r0,1,"Check"); _h(ws,r0,2,"Threshold")
    for d in range(n_days): _h(ws,r0,3+d,f"Day {d+1}")
    r0+=1

    # ATM error check
    _s(ws,r0,1,"ATM IV error < 0.001bps (zero by construction)",align='left')
    _s(ws,r0,2,"< 0.001bps",align='center',bg=ALT)
    for d in range(n_days):
        all_pass=all(abs((results[d][name]['mdl_ivs'][ATM_IDX]-results[d][name]['mkt_ivs'][ATM_IDX])*10000)<0.001
                     for name,_,__ in tenors if results[d].get(name))
        _s(ws,r0,3+d,"PASS" if all_pass else "FAIL",bold=True,fg=WHITE,
           bg=PASS_BG if all_pass else FAIL_BG,align='center')
    r0+=1

    # rho < 0 check (USD convention)
    _s(ws,r0,1,"ρ < 0 for all tenors (USD: negative skew expected)",align='left',bg=ALT)
    _s(ws,r0,2,"ρ < 0",align='center',bg=ALT)
    for d in range(n_days):
        all_neg=all(results[d][name]['rho']<0 for name,_,__ in tenors if results[d].get(name))
        _s(ws,r0,3+d,"PASS" if all_neg else "WARN",bold=True,fg=WHITE,
           bg=PASS_BG if all_neg else WARN_BG,align='center')
    r0+=1

    # nu in range
    _s(ws,r0,1,"ν in [0.05, 1.5] (reasonable range for USD swaptions)",align='left')
    _s(ws,r0,2,"[0.05,1.5]",align='center',bg=ALT)
    for d in range(n_days):
        all_ok=all(0.05<=results[d][name]['nu']<=1.5 for name,_,__ in tenors if results[d].get(name))
        _s(ws,r0,3+d,"PASS" if all_ok else "WARN",bold=True,fg=WHITE,
           bg=PASS_BG if all_ok else WARN_BG,align='center')
    r0+=1

    # Cost J below threshold
    cost_thresh=1e-6
    _s(ws,r0,1,f"Cost J < {cost_thresh:.0e} for all tenors (good IV fit)",align='left',bg=ALT)
    _s(ws,r0,2,f"<{cost_thresh:.0e}",align='center',bg=ALT)
    for d in range(n_days):
        all_ok=all(results[d][name]['cost_J']<cost_thresh for name,_,__ in tenors if results[d].get(name))
        _s(ws,r0,3+d,"PASS" if all_ok else "WARN",bold=True,fg=WHITE,
           bg=PASS_BG if all_ok else WARN_BG,align='center')
    r0+=1

    _note(ws,r0+1,1,
          "PASS=green. WARN=orange (investigate but not necessarily wrong). FAIL=red (action needed). "
          "ATM error should always be ~0 by construction. ρ<0 is USD convention but may differ in stress. "
          f"ν range [0.05,1.5] is a soft sanity range. Cost J<{cost_thresh:.0e} indicates good smile fit.",
          n_days+6)
    ws.freeze_panes="A3"

    # ── 8. Model IVs ─────────────────────────────────────────────────────────
    ws=wb.create_sheet("8_Model_IVs_bps")
    _set_grid_widths(ws,N_STR)
    _h(ws,1,1,"DERIVED: SABR Model Implied Normal Vols (bps). σ_SABR(K)=σ₀·(z/χ(z))·correction using calibrated {σ₀,ρ,ν}. ATM error=0.0000bps by construction (σ₀ enforced each L-M iteration).",bg=HDR_GREEN,sz=9,wrap=True)
    ws.merge_cells(f'A1:{get_column_letter(N_STR+1)}1'); ws.row_dimensions[1].height=30
    write_stacked_grid(ws,results,tenors,n_days,
                       val_fn=lambda r,s: round(r['mdl_ivs'][s]*10000,2),fmt='0.00',day_bg=HDR_GREEN)
    # Legend
    _write_legend(ws, 2, N_STR+3, "COLOUR LEGEND",
        [(AMBER_BG, BLACK_, "Amber", "ATM strike model IV. Error vs market is exactly 0.0000bps by construction — σ₀ is updated every L-M iteration to enforce σ_SABR(ATM)=σ_mkt_ATM. Non-ATM cells have no colour coding; errors are shown in Tab 9."),
        ], col_widths=(10,45))
    ws.freeze_panes="A2"

    # ── 9. IV Fit Error ───────────────────────────────────────────────────────
    ws=wb.create_sheet("9_IV_Fit_Error_bps")
    _set_grid_widths(ws,N_STR)
    _h(ws,1,1,"ERROR (Vol Space, bps): Model IV − Market IV = σ_SABR(K)−σ_mkt(K) in bps. ATM=amber (0 by construction). Green=|err|<0.5bps. Amber=<1bps. Red=>1bps. Large wing errors acceptable if price error is small (vega≈0).",bg=HDR_TEAL,sz=9,wrap=True)
    ws.merge_cells(f'A1:{get_column_letter(N_STR+1)}1'); ws.row_dimensions[1].height=30
    def iv_ebg(v): return GREEN_BG if abs(v)<0.5 else (AMBER_BG if abs(v)<1.0 else FAIL_BG)
    write_stacked_grid(ws,results,tenors,n_days,
                       val_fn=lambda r,s: round((r['mdl_ivs'][s]-r['mkt_ivs'][s])*10000,4),
                       fmt='0.0000',bg_fn=iv_ebg,day_bg=HDR_TEAL)
    # Legend
    _write_legend(ws, 2, N_STR+3, "COLOUR LEGEND",
        [(AMBER_BG, BLACK_, "Amber",  "ATM strike. Error=0.0000bps by construction (σ₀ enforced). Also used for |error| between 0.5 and 1.0bps on non-ATM strikes."),
         (GREEN_BG, BLACK_, "Green",  "|Model IV − Market IV| < 0.5bps. Excellent fit. SABR closely matches the market smile at this strike."),
         (AMBER_BG, BLACK_, "Amber*", "|error| between 0.5 and 1.0bps. Acceptable fit. Minor SABR approximation error or small market noise."),
         (FAIL_BG,  WHITE,  "Red",    "|error| > 1.0bps. Poor fit at this strike. Check: (1) if vega is near-zero (wing) this is economically harmless — see Tab 10 price error; (2) if near-ATM, investigate calibration."),
        ], col_widths=(10,45))
    ws.freeze_panes="A2"

    # ── 10. Price Round-Trip ──────────────────────────────────────────────────
    ws=wb.create_sheet("10_Price_RoundTrip")
    GAP=2; G2=N_STR+2+GAP; G3=G2+N_STR+1+GAP
    total_w=G3+N_STR

    ws.column_dimensions['A'].width=9
    for s in range(N_STR): ws.column_dimensions[get_column_letter(2+s)].width=9
    for g in range(GAP): ws.column_dimensions[get_column_letter(N_STR+2+g)].width=3
    ws.column_dimensions[get_column_letter(G2)].width=9
    for s in range(N_STR): ws.column_dimensions[get_column_letter(G2+1+s)].width=9
    for g in range(GAP): ws.column_dimensions[get_column_letter(G2+N_STR+1+g)].width=3
    ws.column_dimensions[get_column_letter(G3)].width=9
    for s in range(N_STR): ws.column_dimensions[get_column_letter(G3+1+s)].width=9

    _h(ws,1,1,f"PRICE ROUND-TRIP: Re-generated prices from {{σ₀,ρ,ν}} vs Input Totem prices. "
              f"Grid 1: Input market prices (fraction of notional). "
              f"Grid 2: Error in bps of notional = (Re-gen−Input)×10000. "
              f"Grid 3: Error as % of input price. N/M shown when price<{NM_PRICE_THRESH*10000:.1f}bps (near-zero price, % meaningless).",
       bg=HDR_NAVY,sz=9,wrap=True)
    ws.merge_cells(f'A1:{get_column_letter(total_w)}1'); ws.row_dimensions[1].height=36

    def price_bp_bg(v): return GREEN_BG if abs(v)<0.1 else (AMBER_BG if abs(v)<0.5 else FAIL_BG)
    def price_pct_bg(v):
        if v=="N/M": return NM_BG
        return GREEN_BG if abs(v)<0.5 else (AMBER_BG if abs(v)<1.0 else FAIL_BG)

    row=2
    for d in range(n_days):
        # Day header
        _h(ws,row,1,f"DAY {d+1}",bg=HDR_TEAL)
        ws.merge_cells(f'A{row}:{get_column_letter(total_w)}{row}'); row+=1

        # Grid sub-headers
        _h(ws,row,1,"Input Market Price (fraction of notional)",bg=HDR_NAVY)
        ws.merge_cells(f'A{row}:{get_column_letter(N_STR+1)}{row}')
        _h(ws,row,G2,"Price Error (bps of notional)",bg=HDR_GREEN)
        ws.merge_cells(f'{get_column_letter(G2)}{row}:{get_column_letter(G2+N_STR)}{row}')
        _h(ws,row,G3,"Price Error (% of input price)",bg=HDR_TEAL)
        ws.merge_cells(f'{get_column_letter(G3)}{row}:{get_column_letter(total_w)}{row}')
        row+=1

        # Strike headers — all 3 grids
        for gc in [1,G2,G3]:
            _h(ws,row,gc,"Tenor",bg=HDR_NAVY)
            for s,o in enumerate(OFFSETS): _h(ws,row,gc+1+s,_offset_label(o),bg=HDR_NAVY)
        row+=1

        # Data rows
        for i,(name,_,__) in enumerate(tenors):
            bg=ALT if i%2==0 else None
            r=results[d].get(name)
            for gc in [1,G2,G3]: _s(ws,row,gc,name,bold=True,bg=bg)
            for s,o in enumerate(OFFSETS):
                is_atm=abs(o)<1e-6
                if r:
                    inp_p=r['input_prices'][s]
                    reg_p=r['regen_prices'][s]
                    err_bp=(reg_p-inp_p)*10000
                    if abs(inp_p)<NM_PRICE_THRESH:
                        err_pct="N/M"
                    else:
                        err_pct=round((reg_p-inp_p)/inp_p*100,4) if abs(inp_p)>1e-10 else "N/M"
                else:
                    inp_p=0; err_bp=0; err_pct="N/M"

                # Grid 1: market price
                _i(ws,row,2+s,round(inp_p,7),fmt='0.0000000',
                   bg=AMBER_BG if is_atm else bg)
                # Grid 2: bps error
                _c(ws,row,G2+1+s,round(err_bp,4),fmt='0.0000',
                   bg=AMBER_BG if is_atm else price_bp_bg(err_bp))
                # Grid 3: % error — always show number, append (N/M) in brackets if not meaningful
                if is_atm:
                    _s(ws,row,G3+1+s,"0.0000%",align='center',bg=AMBER_BG)
                else:
                    raw_pct=round((reg_p-inp_p)/inp_p*100,4) if r and abs(inp_p)>1e-10 else 0
                    is_nm=abs(inp_p)<NM_PRICE_THRESH if r else True
                    display=f"{raw_pct:.4f}% (N/M)" if is_nm else f"{raw_pct:.4f}%"
                    cbg=NM_BG if is_nm else price_pct_bg(raw_pct)
                    _s(ws,row,G3+1+s,display,align='center',bg=cbg,
                       italic=is_nm,fg=BLACK_)
            row+=1

        # RMSE row
        for gc in [1,G2,G3]: _s(ws,row,gc,"RMSE",bold=True,bg=HDR_NAVY,fg=WHITE)
        _note(ws,row,2,
              "RMSE = √(mean of squared price errors across all tenors) for each strike column. "
              "Example: RMSE=0.05bps at +100bps means the typical calibration price error at that strike "
              "is 0.05bps of notional across all tenors. Useful single summary per strike — "
              "lower RMSE = better fit. ATM RMSE≈0 by construction.",
              N_STR+1,bg=AMBER_BG)
        for s,o in enumerate(OFFSETS):
            is_atm=abs(o)<1e-6
            bps_errs=[]; pct_errs=[]
            for name,_,__ in tenors:
                r=results[d].get(name)
                if r:
                    inp_p=r['input_prices'][s]; reg_p=r['regen_prices'][s]
                    bps_errs.append(((reg_p-inp_p)*10000)**2)
                    if abs(inp_p)>=NM_PRICE_THRESH and abs(inp_p)>1e-10:
                        pct_errs.append(((reg_p-inp_p)/inp_p*100)**2)
            rmse_bp=round(np.sqrt(np.mean(bps_errs)),4) if bps_errs else 0
            rmse_pct=round(np.sqrt(np.mean(pct_errs)),4) if pct_errs else None
            _c(ws,row,G2+1+s,rmse_bp,fmt='0.0000',bg=AMBER_BG if is_atm else price_bp_bg(rmse_bp))
            if is_atm:
                _s(ws,row,G3+1+s,"0.0000% (ATM)",align='center',bg=AMBER_BG)
            elif rmse_pct is None:
                _s(ws,row,G3+1+s,"N/M",align='center',bg=NM_BG,italic=True)
            else:
                _c(ws,row,G3+1+s,rmse_pct,fmt='0.0000"%"',bg=price_pct_bg(rmse_pct))
        row+=1
        _blank_row(ws,row,total_w); row+=1

    # Legend on right side of Grid 3
    legend_col=G3+N_STR+2
    ws.column_dimensions[get_column_letter(legend_col)].width=8
    ws.column_dimensions[get_column_letter(legend_col+1)].width=38
    _h(ws,2,legend_col,"Colour",bg=HDR_NAVY); _h(ws,2,legend_col+1,"Meaning (% Error Grid)",bg=HDR_NAVY)
    leg_items=[
        (GREEN_BG, BLACK_,  "Green",   "|% error| < 0.5% — tight fit, reliable"),
        (AMBER_BG, BLACK_,  "Amber",   "|% error| 0.5%–1.0% — acceptable, monitor"),
        (FAIL_BG,  WHITE,   "Red",     "|% error| > 1.0% — poor fit, investigate"),
        (NM_BG,    BLACK_,  "Grey",    "N/M — Not Meaningful. Option price below "
                                        f"{NM_PRICE_THRESH*10000:.1f}bps of notional. "
                                        "Option is near-worthless; % error is mathematically "
                                        "large but economically irrelevant (vega≈0, price≈0). "
                                        "Look at bps error grid instead."),
        (AMBER_BG, BLACK_,  "Amber*",  "ATM cells always amber — % error is 0 by construction "
                                        "but shown as amber to flag ATM row visually."),
    ]
    for li,(bg,fg,lbl,desc) in enumerate(leg_items,3):
        _s(ws,li,legend_col,lbl,bold=True,fg=fg,bg=bg,align='center')
        _s(ws,li,legend_col+1,desc,align='left',wrap=True,bg=GREEN_BG if bg==GREEN_BG else None)
        ws.row_dimensions[li].height=40 if "near-worthless" in desc else 20

    _note(ws,row,1,
          f"Grid 1 (blue)=input Totem prices (fraction of notional). "
          f"Grid 2: (Re-gen price−Input price)×10000 in bps. "
          f"Grid 3: (Re-gen−Input)/Input×100% with (N/M) shown in brackets where price<{NM_PRICE_THRESH*10000:.1f}bps. "
          "ATM=amber (0 by construction). See colour legend on right of Grid 3. "
          "RMSE=Root Mean Square Error across all tenors per strike — single summary of fit quality.",
          total_w)
    ws.freeze_panes="A2"

    # ── 11. IV Smile All Days ─────────────────────────────────────────────────
    ws=wb.create_sheet("11_IV_Smile_AllDays")
    GAP=2; col2=N_STR+2+GAP; total_w2=col2+N_STR
    ws.column_dimensions['A'].width=9
    for s in range(N_STR): ws.column_dimensions[get_column_letter(2+s)].width=9
    for g in range(GAP): ws.column_dimensions[get_column_letter(N_STR+2+g)].width=3
    ws.column_dimensions[get_column_letter(col2)].width=9
    for s in range(N_STR): ws.column_dimensions[get_column_letter(col2+1+s)].width=9

    _h(ws,1,1,"IV SMILE: Market IV (bps) vs SABR Model IV (bps). Left=Market IV (Bachelier inversion). Right=SABR Model IV (calibrated). Skew: ρ<0→lower strikes have higher IV. Smile: ν>0→wings above linear extrapolation.",bg=HDR_NAVY,sz=9,wrap=True)
    ws.merge_cells(f'A1:{get_column_letter(total_w2)}1'); ws.row_dimensions[1].height=30

    row=2
    for d in range(n_days):
        _h(ws,row,1,f"DAY {d+1}",bg=HDR_TEAL)
        ws.merge_cells(f'A{row}:{get_column_letter(total_w2)}{row}'); row+=1
        _h(ws,row,1,"Market IV (bps)",bg=HDR_TEAL)
        ws.merge_cells(f'A{row}:{get_column_letter(N_STR+1)}{row}')
        _h(ws,row,col2,"SABR Model IV (bps)",bg=HDR_GREEN)
        ws.merge_cells(f'{get_column_letter(col2)}{row}:{get_column_letter(total_w2)}{row}'); row+=1
        for gc in [1,col2]:
            _h(ws,row,gc,"Tenor",bg=HDR_NAVY)
            for s,o in enumerate(OFFSETS): _h(ws,row,gc+1+s,_offset_label(o),bg=HDR_NAVY)
        row+=1
        for i,(name,_,__) in enumerate(tenors):
            bg=ALT if i%2==0 else None
            r=results[d].get(name)
            for gc in [1,col2]: _s(ws,row,gc,name,bold=True,bg=bg)
            for s,o in enumerate(OFFSETS):
                is_atm=abs(o)<1e-6
                mkt=round(r['mkt_ivs'][s]*10000,2) if r else 0
                mdl=round(r['mdl_ivs'][s]*10000,2) if r else 0
                _c(ws,row,2+s,mkt,fmt='0.00',bg=AMBER_BG if is_atm else bg)
                _c(ws,row,col2+1+s,mdl,fmt='0.00',bg=AMBER_BG if is_atm else bg)
            row+=1
        _blank_row(ws,row,total_w2); row+=1

    # Charts Day 1
    chart_row=row+1
    _h(ws,row,1,"CHART DATA — Day 1",bg=HDR_NAVY)
    ws.merge_cells(f'A{row}:{get_column_letter(N_TEN*2+1)}{row}'); row+=1
    _h(ws,row,1,"Offset(bps)")
    col=2
    for name,_,__ in tenors:
        _h(ws,row,col,f"{name} Mkt",bg=HDR_TEAL); _h(ws,row,col+1,f"{name} Mdl",bg=HDR_GREEN); col+=2
    row+=1; cds=row
    for j,o in enumerate(OFFSETS):
        bg=ALT if j%2==1 else None
        _c(ws,row,1,int(o*10000),fmt='0',bg=bg); col=2
        for name,_,__ in tenors:
            r=results[0].get(name)
            _c(ws,row,col,  round(r['mkt_ivs'][j]*10000,2) if r else 0,fmt='0.00',bg=bg)
            _c(ws,row,col+1,round(r['mdl_ivs'][j]*10000,2) if r else 0,fmt='0.00',bg=bg)
            col+=2
        row+=1
    cde=row-1
    anchors=["S2","S18","S34","S50"]
    for idx,(t_s,t_e) in enumerate([(i,min(i+3,N_TEN)) for i in range(0,N_TEN,3)]):
        if idx>=len(anchors): break
        ch=LineChart()
        ch.title=f"IV Smile Day 1 — {', '.join(t[0] for t in tenors[t_s:t_e])}"
        ch.style=10; ch.y_axis.title="Normal Vol (bps)"; ch.x_axis.title="Strike Offset (bps)"
        ch.width=22; ch.height=13
        for t_idx in range(t_s,t_e):
            mc=2+t_idx*2; dc=3+t_idx*2
            ch.add_data(Reference(ws,min_col=mc,max_col=mc,min_row=cds-1,max_row=cde),titles_from_data=True)
            ch.add_data(Reference(ws,min_col=dc,max_col=dc,min_row=cds-1,max_row=cde),titles_from_data=True)
        ch.set_categories(Reference(ws,min_col=1,min_row=cds,max_row=cde))
        ws.add_chart(ch,anchors[idx])
    ws.freeze_panes="A2"

    wb.save(out_path)
    print(f"\nOutput saved: {out_path}")

# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    print("="*65)
    print("  SABR VOL SURFACE CALIBRATION — USD SWAPTIONS (β=0)")
    print("="*65)
    print(f"\nSettings: ρ_init={INITIAL_RHO} ν_init={INITIAL_NU} "
          f"ρ∈[{RHO_LOWER},{RHO_UPPER}] ν∈[{NU_LOWER},{NU_UPPER}] "
          f"stab: ρ<{RHO_STAB_THRESH} ν<{NU_STAB_THRESH} "
          f"N/M threshold={NM_PRICE_THRESH*10000:.1f}bps")
    while True:
        try:
            n_days=int(input("\nNumber of trading days: "))
            if n_days>=1: break
            print("  Must be >=1")
        except ValueError: print("  Enter a whole number")
    while True:
        path=input("Path to input Excel file: ").strip()
        if os.path.exists(path): break
        print(f"  Not found: {path}")
    default_out="sabr_output.xlsx"
    out=input(f"Output file name [default: {default_out}]: ").strip()
    if not out: out=default_out
    if not out.endswith(".xlsx"): out+=".xlsx"

    try:
        tenors,atm_by_day,prices_by_day=read_input(path,n_days)
    except Exception as e:
        print(f"\nERROR reading input: {e}"); sys.exit(1)

    print("\nRunning input sanity checks...")
    sanity_checks=run_input_sanity(tenors,atm_by_day,prices_by_day,n_days)
    for chk in sanity_checks:
        status="PASS" if chk['passed'] else f"FAIL ({len(chk['failures'])} issues)"
        print(f"  {chk['name']}: {status}")

    print(f"\nCalibrating ({len(tenors)} tenors × {n_days} days)...")
    results=run_calibration(tenors,atm_by_day,prices_by_day,n_days)

    print("\nWriting output...")
    try:
        write_output(results,tenors,atm_by_day,prices_by_day,n_days,sanity_checks,out)
    except Exception as e:
        print(f"\nERROR writing output: {e}"); sys.exit(1)
    print(f"Done. Open {out}")

if __name__=="__main__":
    main()
